"""
3 DE OUTUBRO 01:32
"""

import openpyxl
import random

# carregando a planilha do Excel
book = openpyxl.load_workbook('candidatos.xlsx')


# imprimir todos os candidatos(separados por cargo)
def imprimir_todos(lista_candidatos):
    page = book[lista_candidatos]
    for rows in page.iter_rows():
        print(rows[0].value, rows[1].value)


# imprimir todos os candidatos a presidente
def imprimir_presidente():
    page = book['presidente']
    for rows in page.iter_rows():
        print(rows[0].value, rows[1].value)


# sortear um candidato a presidente
def sortear_presidente():
    page = book['presidente']
    random_index = random.randint(1, page.max_row)
    for rows in page.iter_rows(min_row=random_index, max_row=random_index):
        print('Para presidente:', rows[0].value)
        print('Número do candidato:',  rows[1].value)
        print('Partido do candidato:',  rows[2].value)
        print('---------------------------------------------------')

# imprimir todos os candidatos a senador
def imprimir_senador():
    page = book['senador']
    for rows in page.iter_rows():
        print(rows[0].value, rows[1].value)


# sortear um candidato a senador
def sortear_senador():
    page = book['senador']
    random_index = random.randint(1, page.max_row)
    for rows in page.iter_rows(min_row=random_index, max_row=random_index):
        print('Para senador:', rows[0].value)
        print('Número do candidato:', rows[1].value)
        print('Partido do candidato:',  rows[2].value)
        print('---------------------------------------------------')

# imprimir todos os candidatos a governador
def imprimir_governador():
    page = book['governador']
    for rows in page.iter_rows():
        print(rows[0].value, rows[1].value)


# sortear um candidato a governador
def sortear_governador():
    page = book['governador']
    random_index = random.randint(1, page.max_row)
    for rows in page.iter_rows(min_row=random_index, max_row=random_index):
        print('Para governador:',rows[0].value)
        print('Núemro do candidato:', rows[1].value)
        print('Partido do candidato:',  rows[2].value)
        print('---------------------------------------------------')

# imprimir todos os candidatos a deputado federal
def imprimir_dep_federal():
    page = book['dep-federal']
    for rows in page.iter_rows():
        print(rows[0].value, rows[1].value)


# sortear um candidato a dep. federal
def sortear_dep_federal():
    page = book['dep-federal']
    random_index = random.randint(1, page.max_row)
    for rows in page.iter_rows(min_row=random_index, max_row=random_index):
        print('Para deputado federal:',rows[0].value)
        print('Núemro do candidato:', rows[1].value)
        print('Partido do candidato:',  rows[2].value)
        print('---------------------------------------------------')


# imprimir todos os candidatos a dep. distrital
def imprimir_dep_distrital():
    page = book['dep-distrital']
    for rows in page.iter_rows():
        print(rows[0].value, rows[1].value)


# sortear um candidato a dep. distrital
def sortear_dep_distrital():
    page = book['dep-distrital']
    random_index = random.randint(1, page.max_row)
    for rows in page.iter_rows(min_row=random_index, max_row=random_index):
        print('Para deputado distrital:',rows[0].value)
        print('Núemro do candidato:', rows[1].value)
        print('Partido do candidato:',  rows[2].value)
        print('---------------------------------------------------')



def escolher_candidatos():
    sortear_dep_federal()
    sortear_dep_distrital()
    sortear_senador()
    sortear_governador()
    sortear_presidente()
